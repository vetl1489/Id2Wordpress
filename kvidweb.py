#!/usr/bin/python
# -*- coding: UTF-8 -*-

# KVIDWEB v. 1.0.1

#############################################################
# Перед использованием установить:
# Библиотеку для работы с WordPress по протоколу XML-RPC
# $ pip install python-wordpress-xmlrpc
# Библиотеку для работы с таблицами MS Excel 2007 и старше
# $ pip install openpyxl
# Библиотеку для раскрашивания консоли
# $ pip install colorama
#############################################################

from __future__ import print_function
from colorama import init, Fore
import os
import re
import sys
import time
import fnmatch
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Style, Font, PatternFill
from urllib.parse import urljoin
import argparse
import json

#############################################################
# Параметры
#############################################################


file_settings = os.path.join( os.path.dirname(__file__), 'kv_settings.json' )
settings = json.load( open(file_settings) )

# файл для обработки по умолчанию
my_html = settings['my_html']   
# папка для вывода статей по умолчанию
my_dir = settings['my_dir']                
# формат выходных файлов
type_file = settings['type_file']            
# имя файла без заголовка
default_name = settings['default_name']
table = settings['table']

# объявляем массив кортежей поиска-замены
regexps = [
    # удаляем <!DOCTYPE html>
    ( r'<[!]DOCTYPE html>', r'' ),
    # удаляем открывающие и закрывающие div|html|head|meta|link|body|img
    ( r'<[/]?(div|html|head|meta|link|body|img)(.+)?>', r'' ),
    # удаляем title с его содержимым
    ( r'<title>.+</title>', r'' ),
    # удаляем классы в заголовке
    ( r'(?<=h1) class=["].+["]', r'' ),
    # удаляем класс kill
    ( r'(\sstyle=["](\w+?)?["])?(\sclass=["]kill["])?', r'' ),
    # удаляем тег br
    ( r'<br(\s[/])?>', r'' ),
    # удаляем забытые классы с кириллицей
    ( r'(?<=p)(\sclass=["][А-я0-9].+["])(?=><)', r'' ),
    ( r'(?<=strong)(\sclass=["][А-я0-9].+?["])(?=>)', r'' ),
    ( r'(?<=li)(\sclass=["][А-я0-9].+?["])(?=>)', r'' ),
    # удаляем повторяющиеся strong
    ( r'(<[/]strong><strong>)', r'' ),
    # удаляем пустые ссылки
    ( r'(<a\s.+?><[/]a>)', r'' ),
    # удаляем табуляцию перед тегами
    ( r'(\t+(?=<))', r'' ),
    # удаляем табуляцию в пустых областях
    ( r'(\t+(?=\n))', r'' ),
    # удаляем переводя каретки в начале
    ( r'(^(\r?\n)+)', r'' ),
    # удаляем спаны с классами из DoTextOK
    ( r'(<span class=["](myHorizontalScaleForEmDashSpace|mySupChars|mySubChars|myBoldChars|myItalicChars|myBoldItalicChars|mySpecialChars|mySupSpecialChars|mySubSpecialChars|mySymbolChars|mySubSymbolChars|mySupSymbolChars)["]>)(.+?)(<[/]span>)', r'\3' ),
    # удаляем пустые классы
    ( r'(<span>)(.+?)(</span>)', r'\g<2>' ),
    # прибиваем подпись к статье
    ( r'(<[/]p>)(\r?\n?)*(<p\sclass=["]signed["]>)', r'\1\n\3' ),
    # делаем отбивки между статьями
    ( r'(<[/]p>)(\r?\n?)+(<h1>)', r'\1\n\n\n\n\3' ), 
    # меняем неразрывные пробелы на правильные
    ( r'&#160;', r'&nbsp;' ),
    # убираем мусор в конце текста
    ( r'(\n)*\Z', r'' ),
    
    # обрабатываем стихи
    ( r'(<p\sclass=["]lyric["]>.+?)(</p>)(\r?\n?\s+?)(<p\sclass=["]lyric["]>)', r'\1<br>\3' ),
    ( r'(<br>\r?\n?\s?.+?)(</p>)(\r?\n?\s?)(<p\sclass=["]lyric["]>)', r'\1<br>\3' ),

    # делаем первый абзац лидом, там где его нет
    ( r'(<[/]h1>(\r?\n?)*<p)(>)', r'\g<1> class="big"\g<3>' ),
    ( r'(</h1>\r?\n?\s?<p\sclass=["]lyric["]>)((.+?<(br|/p)>\r?\n?\s?)+?)(<p)', r'\1\2\5 class="big"' ),
    
    # Убираем буквицу
    ( r'(<span\sclass=["]dropcap["]>)(.+?)(</span>)', r'\g<2>' ),
    # отбиваем заголовок от статьи
    ( r'(<[/]h1>\n)', r'\1\n' ),
    # меняем короткое тире на длинное
    ( r'(;|>|\s)(\u2013)', r'\1&mdash;' ),
    ( r'(\d)(\u2013)(\d)', r'\1&ndash;\3' ),
    # убираем отбивки между тегами заголовков h1 и сливаем их
    ( r'<[/]h1>(\r?\n?)*<h1>', r'<br>' ),
    # убирем теги h1
    ( r'<[/]?h1>', r'' ),
    
    # добавляем тег читать далее
    # ( r'\n\n<p.+>.+?</p>\n', r'\g<0><!--more-->\n' ),
    ( r'(<p\sclass=["]big["]>.+?</p>\r?\n?\s?)', r'\g<0><!--more-->\n' ),

    # объединяем две идущие подряд подписи в одну
    ( r'(<p class=["]signed["]>)(.+?)(</p>)(\r?\n?\s?)(<p class=["]signed["]>)', r'\1\2 ' ),
    ( r'(<p class=["]signed["]>)(.+?)(</p>)(\r?\n?\s?)(<p class=["]signed["]>)', r'\1\2 ' ),
    ( r'(<p class=["]signed["]>)(.+?)(</p>)(\r?\n?\s?)(<p class=["]signed["]>)', r'\1\2 ' ),
    
]

######################################################
# помощь
######################################################
myHelp = """
Программа для обработки файла HTML со статьями, выведенными из Adobe InDesign для блога на WordPress. По-умолчанию статьи выводятся в папку "WEB", файл для обработки по-умолчанию должен называться "all.html".

Папку и файл можно задать другими, передав их 
в качестве параметров командной строки.
"""

######################################################
# Функции
######################################################

# выводим сообщение
def report (message, error = False) :
    if error :
        init()
        print (Fore.RED + 'ERROR: ' + message + Fore.RESET)
        input('Нажмите [Enter] для выхода')
        sys.exit()
    else :
        init()
        print (Fore.YELLOW + message + Fore.RESET)
        input('Нажмите [Enter] для выхода')
        sys.exit()

# исправляем названия файлов статей
def fix_name (name) :
    list_re = [
        ( r'&mdash;', r'-' ),
        ( r'&ndash;', r'-' ),
        ( r'&nbsp;', r' ' ),
        ( r'\u00AB', '' ),
        ( r'\u00BB', '' ),
        ( r'\u201E', '' ),
        ( r'\u201C', '' ),
        ( r'[:;]', r'' ),
        ( r'\u2026', r'' ),
        ( r'<.+?>', r' ' ),
        ( r'\s{2,}', r' ' ),
    ]
    for r in range (len(list_re)):
        reOk = re.search(list_re[r][0], name)
        if reOk :
            name = re.sub(list_re[r][0], list_re[r][1], name)
    return name

# тестовая функция сохранения файла с каким-либо контентом
def out_file (content, filename) :
    b = open(filename + '.html', 'w', encoding='utf-8')
    b.write(str(content))
    b.close()

# парсер параметров командной строки
def createParser ():
    parser = argparse.ArgumentParser(
        prog = 'kvidweb',
        description = myHelp,
        epilog = '(c) vetl1489, 2015.',
        # add_help = False,
        )
    parser.add_argument ('-html', '--html', nargs='?', default=my_html, help='Имя html файла для обработки', metavar='ФАЙЛ HTML' )
    parser.add_argument ('-f', '--folder', nargs='?', default=my_dir, help='Имя папки с выходными файлами', metavar='ПАПКА' )
    return parser

# выдаем корректное имя файла
def set_name(count, name):
    length = 252
    pref = ' - '
    zero = '0' if count < 9 else ''
    dir_to = os.path.abspath(my_dir)
    full_length = len(dir_to) + len(zero) + len(str(count)) + len(pref) + len(name) + len(type_file)
    
    if length - (len(dir_to) - len(zero) + len(str(count)) + len(pref) + len(type_file)) <= 1:
        report('Путь к файлу длиннее разрешенных в Windows 256 символов', error=True)
    filename = name
    if full_length > length:
        filename = name[0:length-len(dir_to)-len(type_file)]

    return os.path.join(dir_to, zero + str(count + 1) + pref + filename + type_file)

######################################################
# Программа
######################################################

parser = createParser()
args = parser.parse_args(sys.argv[1:])

my_html = args.html
my_dir = args.folder

report ('Начали...')

# проверяем наличие файла и открываем его
if os.path.exists(my_html):
    f = open (my_html, 'r', encoding='utf-8')
    content = f.read()
    f.close()
else:
    report ('Отсутсвует файл для обработки', error = True)

# обрабатываем регулярками
for i in regexps:
    find = re.compile(i[0])
    replace = i[1]
    if re.search(find, content):
        content = re.sub(find, replace, content)

# разбиваем текст на массив со статьями
arrArticles = re.split(r'\n{3,}', content)

# вынимаем названия статей и записываем в список
filename = []
count = 1
for item in arrArticles:
    match = re.search(r'^.+(?=\n{2})', item)
    if match:
        na = match.group()
    else:
        na = default_name + str(count)
        count+=1
    filename.append( fix_name(str(na)) )

# проверяем существование папки куда складываются статьи
if not os.path.exists(my_dir) :
    os.mkdir(my_dir)

# записываем файлы статей в папку
for c, art in enumerate(filename):
    nameArticle = set_name(c, art)
    with open(nameArticle, 'w', encoding='utf-8') as article:
        article.write(arrArticles[c])

#######################################################

# Получили список файлов в папке с текстами
arrFiles = []
arrFiles.extend (fnmatch.filter(os.listdir(my_dir), '*' + type_file))

# Создаем документ Excell
wb = Workbook()
ws = wb.active  # Обрщаемся к активному листу
# Названия колонок
fieldnames = ['Статья', 'Рубрики', 'Тэги', 'Миниатюра', 'Опубл.']
# Стиль для шапки
myStyle = Style(font = Font(bold = True), fill = PatternFill(fill_type='solid', start_color='FFD9D9D9', end_color='FF000000')) 
# Записываем и красим шапку
for cnt, fn in enumerate(fieldnames):
     ws.cell(row = 1, column = cnt+1).value = fn
     ws.cell(row = 1, column = cnt+1).style = myStyle

# Заполняем списком статей и статусом опубликованности
# 1 - опубликовано, 0 - не опубликовано
for cn, af in enumerate(arrFiles):
    ws.cell(row=cn+2, column = 1).value = af
    ws.cell(row=cn+2, column = 5).value = 1

# Назначем ширину колонок
ws.column_dimensions['A'].width=45
ws.column_dimensions['B'].width=20
ws.column_dimensions['C'].width=22
ws.column_dimensions['D'].width=30
ws.column_dimensions['E'].width=7

# Сохраняем таблицу
try :
    wb.save(os.path.join(my_dir, table))
except :
    report ('Не могу записать таблицу сопоставлений!', error = True)

report ('Готово!')