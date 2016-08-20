#!/usr/bin/python
# -*- coding: UTF-8 -*-

# KVSend v. 1.0.2

# pip install python-wordpress-xmlrpc
# pip install openpyxl
# pip install colorama


from __future__ import print_function
from colorama import init, Fore
import re
import os
import os.path
import fnmatch
import mimetypes
import time
import sys
import csv
import json
import openpyxl
from openpyxl.cell.cell import Cell
from urllib.parse import urljoin
from wordpress_xmlrpc import Client, WordPressPost
from wordpress_xmlrpc.compat import xmlrpc_client
from wordpress_xmlrpc.methods import media, posts, taxonomies
from wordpress_xmlrpc.methods.posts import GetPosts, NewPost
from wordpress_xmlrpc.methods.taxonomies import GetTaxonomies

#######################################################
# Параметры
#######################################################

file_settings = os.path.join( os.path.dirname(__file__), 'kv_settings.json' )
settings = json.load( open(file_settings) )

# Папка по умолчанию с обработанными статьями 
base_path = settings['base_path']
# Имя таблицы с сопоставлениями
table = settings['table']

# WordPress
my_web_site = settings['my_web_site']
login = settings['login']
password = settings['password']

type_article = settings['type_article']
tmp_file = settings['tmp_file']
type_img = settings['type_img']

# Рубрика по-умолчанию, если не указана в сопоставлениях
default_category = settings['default_category']

# Шаблоны вывода сообщений об отправке
template1 = 'Отправляем материал {item}: "{an}"'
template2 = 'Материал отправлен!\nРубрики: {cat}\nТеги: {tags}\nМиниатюра: {th}\nID поста: {ip}'

# символы для замены, чтобы не сломать консоль
badChars = [
    # тирешки
    (r'\u2013', r'-'),
    (r'\u2014', r'-'),
    # кавычки
    (r'\u00AB', r'"'),
    (r'\u00BB', r'"'),
    (r'\u201E', r'"'),
    (r'\u201C', r'"'),
    # пробелы
    (r'\u00A0', r' '),
    (r'\u202F', r' '),
    (r'\u2009', r' '),
    (r'\u200a', r' '),
    (r'\u2026', r''),
]

#######################################################
# помощь
#######################################################
myHelp = """
kvsend 

Программа для отправки статей на сайт обработанных с помощью kvidweb.
Перед запуском необходимо заполнить таблицу соопоставлений
статей, их таксономий и путей к миниатюрам.
Поле "Опубл." принимает значения:
1 - статья будет опубликована
0 - статья примет статус "Черновик"

Имя таблицы сопоставлений по-умолчанию "Articles.xlsx",
такое, как на выходе из программы kvidewb.
Однако, программа в качестве параметра командной строки 
может принимать иное имя таблицы сопоставлений

send my_file.xlsx
"""

#######################################################
# Функции
#######################################################

# Парсим таксономии
def parse_taxonom (tax) :
    try :
        rawTax = re.split(r'(?<=\w)(\s*[,;]+\s*)|(\s+$)', tax)
        taxList = []
        for item in range(len(rawTax)) :
            if type(rawTax[item]) is str :
                match = re.search(r'\w+', rawTax[item])
                if match :
                    taxList.append (rawTax[item])
        return taxList
    except :
        return None

# Вытаскиваем из файла заголовок и текст статьи и пихаем их список
def get_article (path) :
    art = open (os.path.join(base_path, path), 'r', encoding='utf-8')
    content = art.read()
    article = re.split(r'\n{2}', content)
    return article

# Функция выдачи сообщения
def report (message, error = False) :
    if error :
        print (Fore.RED + 'ERROR: ' + Fore.YELLOW + message + Fore.RESET)
        input('Нажмите [Enter] для выхода\n')
        sys.exit()
    else :
        print (Fore.GREEN + message + Fore.RESET)

# отправляем картинку в блог
def send_img (imageName): 
    with open(imageName, "rb") as img:
        data = {
            'name': os.path.basename(imageName),
            'bits': xmlrpc_client.Binary(img.read()),
            'type': mimetypes.guess_type(imageName)[0],
        }
        # получаем id загруженной картинки чтобы привязать к посту
        response = wp.call(media.UploadFile(data))
        return response['id']

# Выводим человеческий список таксономий
def print_list (tax) :
    s = ''
    if len(tax) < 1 :
        s = 'None'
    else :
        for t in range (len(tax)) :
            s = s + tax[t] + ', '
        s = re.sub(r'()(,\s$)', r' ', s)
    return s

# функция правит текст для вывода в консоль Windows
def kill_bad_chars (text) :
    for i in range (len(badChars)) :
        regexp = re.compile(badChars[i][0])
        replace = badChars[i][1]
        if re.search(regexp, text) :
            text = re.sub(regexp, replace, text)
    return text

# проверка таблицы на ошибки
def verify_table (xl_table) :
    # есть ли лист с нужным именем?
    try : 
        _sheet = xl_table['Sheet']
    except : 
        report ('Неверная таблица сопоставлений', error = True)
    
    # есть ли записи помимо заголовка?
    if len(_sheet.rows) < 2 :
        report ('Отсутствуют записи в таблице', error = True)
    else :
        pass

    # массив с именами файлов с ошибочным статусом публикации
    _err_status = []
    for q in range(len(_sheet.rows)) :
        if q > 0 :
            if (_sheet.cell(row = q+1, column = 5).value == 1) or (_sheet.cell(row = q+1, column = 5).value == 0) :
                pass
            else :
                _err_status.append(_sheet.cell(row = q+1, column = 1).value)

    # проверяем миниатюры
    _err_imgs = [] 
    for c in range(len(_sheet.rows)) :
        if c>0 :
            if _sheet.cell(row = c+1, column = 4).value :
                filename = os.path.normpath(_sheet.cell(row = c+1, column = 4).value)
                if os.path.exists(filename) == False:
                    _err_imgs.append (_sheet.cell(row = c+1, column = 1).value)

    # получили список файлов со статьями из папки
    _files = fnmatch.filter(os.listdir('.'), type_article)

    # получили список статей из таблицы
    _articles = []
    for j in range(len(_sheet.rows)) :
        if j>0 :
            if _sheet.cell(row = j+1, column = 1).value :
                _articles.append(_sheet.cell(row = j+1, column = 1).value)

    _err_files = list(set(_files) - set(_articles))
    _err_articles = list(set(_articles) - set(_files))
    _err_files.sort()
    _err_articles.sort()

    init()
    err_msg = Fore.RED + '\nERROR: ' + Fore.YELLOW
    
    # ключ ошибки
    _err_key = 0 
    if len(_err_status)>0 :
        print ( err_msg + 'Отсутсвует или некорректен статус публикации:' + Fore.RESET )
        for w in _err_status :
            print ( '  ' + w )
        _err_key = _err_key + 1
    if len(_err_files)>0 :
        print ( err_msg + 'В папке имеются файлы, отсутвующие в таблице:' + Fore.RESET )
        for s in _err_files :
            print ( '  ' + s )
        _err_key = _err_key + 1
    if len(_err_articles)>0 :
        print ( err_msg + 'В таблице имеются статьи, отсутвующие в папке:' + Fore.RESET )
        for x in _err_articles :
            print ( '  ' + x )
        _err_key = _err_key + 1
    if len(_err_imgs)>0 :
        print ( err_msg + 'Имеются несоответствия у путей миниатюр и самих файлов:' + Fore.RESET )
        for o in _err_imgs :
            print ( '  ' + o )
        _err_key = _err_key + 1

    if _err_key > 0 :
        input('Нажмите [Enter] для выхода\n')
        sys.exit()

# проверяем текст статьи
def verify_article (table):
    _sheet = table['Sheet']
    
    def _verify (_article, _path) :
        _reg = [r'^\w+', r'<p .+?>.+?<[/]p>']
        if type(_article) is list :
            if len(_article) < 2 :
                if re.search(_reg[0], _article[0]) :
                    return ( str(_path), 0 )
                if re.search(_reg[1], _article[0]) :
                    return ( str(_path), 1 )
            elif len(_article) > 2 :
                return ( str(_path), 2 )
            else :
                return None
        else :
            return ( str(_path), 2 )

    # (name_article, key)
    # key=0  Отсутсвует текст статьи
    # key=1  Отсутсвует заголовок статьи
    # key=2  Нераспознана статья

    # создали массив ошибок
    _err_arr = []
    _files = fnmatch.filter(os.listdir('.'), type_article)
    for _f in range (len(_files)) :
        var = _verify (get_article (_files[_f]), _files[_f])
        if var is not None :
            _err_arr.append (var)

    if len(_err_arr)>0 :
        print ( Fore.RED + '\nERROR:' )
        for z in range (len(_err_arr)) :
            if _err_arr[z][1] == 0 :
                print (Fore.YELLOW + 'Отсутсвует текст или ошибка оформления: ' + Fore.RESET + _err_arr[z][0])
            elif _err_arr[z][1] == 1 :
                print (Fore.YELLOW + 'Отсутсвует заголовок статьи: ' + Fore.RESET + _err_arr[z][0])
            elif _err_arr[z][1] == 2 :
                print (Fore.YELLOW + 'Нераспознана статья: ' + Fore.RESET + _err_arr[z][0])
        # time.sleep(30)
        input('Нажмите [Enter] для выхода\n')
        sys.exit()

# спрашиваем подтверждение об отправлении
def question(ca) :
    ans = input ('\nОтправить материал "' + ca + '"? [Y(да)/N(нет)]: ')
    if ans == 'Y' or ans == 'y' :
        return True
    elif ans == 'N' or ans == 'n' :
        return False
    else :
        print (Fore.GREEN + 'Некорректный ответ' + Fore.RESET)
        return question(ca)

# Читаем и пишем файл текущего состояния выполнения программы
def read_tmp() :
    tmp_arr = []
    with open(tmp_file, newline='') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        for row in spamreader:
            tmp_arr.append(row)
        for st in tmp_arr :
            st[1] = int(st[1])
    return tmp_arr

def write_tmp(tmp_list) :
    with open(tmp_file, 'w', newline='') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
        for row_art in tmp_list :
            spamwriter.writerow(row_art)

# функция изменения статуса опубликованности 
def set_status (arr, count) :
    arr[count-1][1] = 1
    return arr


######################################################
# Сама программа
######################################################
# инициализируем colorama
init()

# проверяем параметры командной строки
if len (sys.argv) > 1 :
    if sys.argv[1] == 'help' or sys.argv[1] == 'h' or sys.argv[1] == '-h':
        report (myHelp)
        input ('Нажмите [Enter] для выхода')
        sys.exit()
    else :
        table = sys.argv[1]
        if os.path.exists(table) == False :
            report ('Отсутствует указанный файл таблицы сопоставлений', error = True)


подключаемся к блогу
try :
    myURL = urljoin(my_web_site, 'xmlrpc.php')
    wp = Client(myURL, login, password)
    rrrrrr = wp.call(GetTaxonomies())
except :
    report ('Неверные настройки подключения или проблемы с сетью', error = True)


# Открываем таблицу сопоставлений
try :
    wb = openpyxl.load_workbook(filename = table)
except : 
    report ('Отсутствует таблица сопоставлений', error = True)
sheet = wb['Sheet']

# проверяем таблицу
verify_table (wb)
# проверяем тексты
verify_article(wb)

# данные для tmp файла
articles = []
if os.path.exists(tmp_file) == False :
    for z in range(len(sheet.rows)) :
        if z>0 :
            if sheet.cell(row = z+1, column = 1).value :
                articles.append([sheet.cell(row = z+1, column = 1).value, 0])
    write_tmp(articles)
else :
    articles = read_tmp()
    # print (articles)

# Основной цикл считывания данных и отправки на сайт
for i in range(len(sheet.rows)) :
    if i>0 :
        currentArticle = []
        # делаем список со значениями из таблицы сопоставлений
        currentArticle.append(get_article(sheet.cell(row = i+1, column = 1).value))
        currentArticle.append(parse_taxonom(sheet.cell(row = i+1, column = 2).value))
        currentArticle.append(parse_taxonom(sheet.cell(row = i+1, column = 3).value))
        currentArticle.append(sheet.cell(row = i+1, column = 4).value)
        currentArticle.append(sheet.cell(row = i+1, column = 5).value)

        # print (Fore.GREEN + currentArticle + Fore.GREEN)
        
        # если не указана категория, назначаем по-умолчанию
        if currentArticle[1] is None :
            currentArticle[1] = [default_category]
        # если не указаны теги, то и ладно
        if currentArticle[2] is None :
            currentArticle[2] = []
        
        # ключ наличия у статьи миниатюры
        thumb = False
        if currentArticle[3] :
            filename = os.path.normpath(currentArticle[3])
            attachment_id = send_img (filename)
            thumb = True
        
        # создаем новый пост с данными из текущего списка
        post = WordPressPost()
        post.title = currentArticle[0][0]
        post.content = currentArticle[0][1]
        # проверяем наличие у поста миниатюры
        if thumb == True :
            post.thumbnail = attachment_id
            m = os.path.basename(filename)
        else :
            m = 'None'
       
        post.terms_names = {
            'category': currentArticle[1],
            'post_tag': currentArticle[2],
        } 

        if currentArticle[4] == 1 :
            post.post_status = 'publish'

            if articles[i-1][1] == 1 : 
                report('\nМатериал "' + kill_bad_chars(articles[i-1][0]) + '" уже отправлен!')
            else :
                status = question(kill_bad_chars(currentArticle[0][0]))
                if status:
                    report ('Отправляем...')
                    time.sleep(1)
                    try:
                        # отправляем пост
                        id_post = wp.call(NewPost(post))
                        report (template2.format(th=m, cat = print_list(currentArticle[1]), tags = print_list(currentArticle[2]), ip = id_post))
                        write_tmp (set_status(articles, i))
                    except :
                        report ('Ошибка отправки! Проверьте соединение с сетью', error=True)
                else :
                     report ('Материал пропущен')
        else :
            print (Fore.RED +  '\nСтатус публикации - 0 (не публиковать). Материал "' + kill_bad_chars(currentArticle[0][0]) + '" пропущен!' + Fore.RESET)

        # чистим список текущей статьи
        currentArticle.clear()
        time.sleep(0.3)

report ('\nГотово!')
input('Нажмите [Enter] для выхода\n')
